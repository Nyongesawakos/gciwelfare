from enum import member
from urllib import request
import threading
from dateutil.relativedelta import relativedelta
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.http import JsonResponse
from django.db.models.functions import Cast
from django.contrib import messages
from django.db.models import CharField
from django.contrib.auth.decorators import login_required,user_passes_test
from django.db.models import Q
from django.contrib.auth.models import User
import openpyxl
from django.views.decorators.cache import never_cache
from openpyxl.utils import get_column_letter
from .models import user, message,Msg
from django.contrib.auth import authenticate, login, logout,update_session_auth_hash
from .models import tips
from .models import room, group,update,cash_expenditure
from .forms import RoomForm, CustomUserCreationForm, MsgForm, message,MessageForm, MpesaTransactionForm,OpeningBalanceForm
from .forms import UpdateForm, cash_expenditureForm
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm
from django.core.exceptions import ValidationError
from django.db.models import Sum
from django.core.paginator import Paginator
from django.db.models import F 
import datetime
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from django.contrib import messages
from .models import MpesaTransaction
from.forms import MpesaForm
import json
from django.http import JsonResponse
from django_daraja.mpesa.core import MpesaClient
from django.utils import timezone
from .forms import BulkMessageForm
from .models import BulkMessage
from .utils import send_bulk_sms
from web import models
from web import models
from datetime import date
from django.utils import timezone
from django.db.models.functions import ExtractMonth, ExtractYear
from datetime import date
from django.db.models import Sum
from django.utils import timezone
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from django.db.models import Sum
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.models import User

from datetime import datetime

from .models import MpesaTransaction, OpeningBalance

# =====================================================
# WELFARE CONTRIBUTION SETTINGS
# =====================================================

MONTHLY_FEE = Decimal("200")
COUPLE_FEE = Decimal("400")

WELFARE_START_DATE = date(2026, 1, 1)


def calculate_welfare_arrears(user):

    # =====================================================
    # GET MONTHLY CONTRIBUTION FEE
    # =====================================================

    monthly_fee = MONTHLY_FEE

    # Try to get a transaction/profile belonging to this user
    transaction = (
        MpesaTransaction.objects
        .filter(
            user=user,
            profile__isnull=False
        )
        .select_related("profile__group")
        .first()
    )

    if transaction and transaction.profile:

        if (
            transaction.profile.group
            and transaction.profile.group.name == "Couple"
        ):
            monthly_fee = COUPLE_FEE


    # =====================================================
    # CURRENT DATE
    # =====================================================

    today = timezone.now().date()

    # We only calculate from the welfare start date
    # up to the current month.
    current_month = today.replace(day=1)


    # =====================================================
    # BUILD ALL WELFARE MONTHS
    # =====================================================

    months = []

    current = WELFARE_START_DATE

    while current <= current_month:

        months.append({
            "month": current.strftime("%B"),
            "year": current.year,
            "required": monthly_fee,
            "paid": Decimal("0.00"),
            "status": "missing",
            "_date": current,
        })

        current += relativedelta(months=1)


    # =====================================================
    # GET SUCCESSFUL PAYMENTS
    # =====================================================

    payments = (
        MpesaTransaction.objects
        .filter(
            user=user,
            status="Success"
        )
        .order_by("created_at", "id")
    )


    # =====================================================
    # ALLOCATE PAYMENTS TO MONTHS
    #
    # Payments are applied from the oldest unpaid month
    # forward.
    #
    # Example:
    #
    # Monthly fee = KSh 2
    # Payment = KSh 6
    #
    # October  = 2
    # November = 2
    # December = 2
    # =====================================================

    for payment in payments:

        payment_amount = Decimal(
            str(payment.amount or "0")
        )

        if payment_amount <= 0:
            continue


        for month in months:

            required = month["required"]

            already_paid = month["paid"]

            remaining = required - already_paid


            # Month is already fully paid
            if remaining <= 0:
                continue


            # Amount that can be applied to this month
            allocation = min(
                payment_amount,
                remaining
            )


            month["paid"] += allocation

            payment_amount -= allocation


            # Update status
            if month["paid"] >= required:

                month["status"] = "paid"

            else:

                month["status"] = "missing"


            # Payment has been completely allocated
            if payment_amount <= 0:
                break


    # =====================================================
    # REMOVE INTERNAL DATE FIELD
    # =====================================================

    for month in months:

        month.pop("_date", None)

        month["required"] = Decimal(
            month["required"]
        ).quantize(Decimal("0.01"))

        month["paid"] = Decimal(
            month["paid"]
        ).quantize(Decimal("0.01"))


    # =====================================================
    # PAID / MISSING MONTHS
    # =====================================================

    paid_months_list = [
        month
        for month in months
        if month["status"] == "paid"
    ]


    missing_months_list = [
        month
        for month in months
        if month["status"] != "paid"
    ]


    # =====================================================
    # COUNTS
    # =====================================================

    paid_months = len(paid_months_list)

    months_missing = len(missing_months_list)


    # =====================================================
    # TOTAL ARREARS
    #
    # If a month was partially paid, only the remaining
    # amount is considered arrears.
    # =====================================================

    amount_due = sum(
        (
            month["required"] - month["paid"]
            for month in missing_months_list
        ),
        Decimal("0.00")
    )


    amount_due = amount_due.quantize(
        Decimal("0.01")
    )


    # =====================================================
    # TOTAL PAID
    # =====================================================

    total_paid_amount = (
        payments.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )


    total_paid_amount = Decimal(
        str(total_paid_amount)
    ).quantize(Decimal("0.01"))


    # =====================================================
    # TOTAL REQUIRED
    # =====================================================

    total_required = sum(
        (
            month["required"]
            for month in months
        ),
        Decimal("0.00")
    )


    # =====================================================
    # RETURN EVERYTHING
    # =====================================================

    return {

        "monthly_fee": monthly_fee,

        "total_required": total_required,

        "total_paid": total_paid_amount,

        "paid_months": paid_months,

        "months_missing": months_missing,

        "amount_due": amount_due,

        "monthly_status": months,

        "paid_months_list": paid_months_list,

        "missing_months_list": missing_months_list,

    }
@require_http_methods(["GET", "POST"])
@login_required(login_url='/login')
@never_cache
def cashflow(request):

    if request.method == "POST":

        form = cash_expenditureForm(request.POST)

        if form.is_valid():

            cash = form.save(commit=False)

            # Save the expenditure
            cash.save()

            return redirect("cashflow")

    else:

        form = cash_expenditureForm()

    # Get all expenditures
    s = cash_expenditure.objects.select_related(
        "member"
    ).all()

    context = {
        "form": form,
        "s": s,
    }

    return render(
        request,
        "web/cash_flow.html",
        context
    )
@require_http_methods(["GET", "POST"])
@login_required(login_url='login')
@never_cache
def add_opening_balance(request, profile_id):

    # =====================================================
    # GET MEMBER PROFILE
    # =====================================================

    profile = get_object_or_404(
        room,
        id=profile_id
    )

    # The member's Django User is stored in room.host.
    # Do NOT use request.user here because request.user
    # will be the admin when the admin is entering the balance.
    member_user = profile.host

    if member_user is None:
        messages.error(
            request,
            "This member profile is not linked to a user account."
        )
        return redirect(
            'mpesa_balance',
            profile_id=profile.id
        )

    # =====================================================
    # HANDLE FORM SUBMISSION
    # =====================================================

    if request.method == 'POST':

        form = OpeningBalanceForm(request.POST)

        if form.is_valid():

            opening_balance = form.save(commit=False)

            # Attach the selected member's profile
            opening_balance.profile = profile

            # Attach the USER belonging to that profile.
            # This is profile.host, NOT request.user.
            opening_balance.user = member_user

            opening_balance.save()

            messages.success(
                request,
                f'Balance brought forward of '
                f'KSh {opening_balance.amount:,.2f} '
                f'was added successfully for {profile}.'
            )

            return redirect(
                'mpesa_balance',
                profile_id=profile.id
            )

    else:

        form = OpeningBalanceForm(
            initial={
                'balance_date': '2024-12-31',
                'description': 'Balance brought forward from 2024',
            }
        )

    # =====================================================
    # RENDER FORM
    # =====================================================

    return render(
        request,
        'web/add_opening_balance.html',
        {
            'form': form,
            'profile': profile,
            'member_user': member_user,
        }
    )

@never_cache
@login_required(login_url='login')
def mpesa_balance(request, profile_id):

    # =====================================================
    # GET MEMBER PROFILE
    # =====================================================

    profile = get_object_or_404(
        room,
        id=profile_id
    )


    # =====================================================
    # OPENING BALANCE / BALANCE BROUGHT FORWARD
    #
    # Amount carried forward from 2024.
    # =====================================================

    opening_balance = OpeningBalance.objects.filter(
        profile=profile
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')


    # =====================================================
    # SUCCESSFUL M-PESA TRANSACTIONS
    #
    # Only transactions belonging to this profile
    # are included.
    # =====================================================

    transactions = MpesaTransaction.objects.filter(
        profile=profile,
        status='Success'
    ).order_by('-created_at')


    # =====================================================
    # TOTAL M-PESA PAYMENTS
    # =====================================================

    mpesa_total = transactions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')


    # =====================================================
    # CURRENT BALANCE
    #
    # Opening Balance + M-Pesa Payments
    # =====================================================

    current_balance = (
        opening_balance +
        mpesa_total
    )


    # =====================================================
    # ALL TRANSACTIONS
    #
    # This allows the template to display the M-Pesa
    # transaction history.
    # =====================================================

    all_transactions = MpesaTransaction.objects.filter(
        profile=profile
    ).order_by('-created_at')


    # =====================================================
    # PAGE CONTEXT
    # =====================================================

    context = {

        # Member
        'profile': profile,

        # Balance information
        'opening_balance': opening_balance,
        'mpesa_total': mpesa_total,
        'current_balance': current_balance,

        # Successful transactions
        'transactions': transactions,

        # All transactions
        'all_transactions': all_transactions,
    }


    # =====================================================
    # RENDER
    # =====================================================

    return render(
        request,
        'web/mpesa_balance.html',
        context
    )
def insert(request):
    form=UpdateForm()
    if request.method == 'POST':
        form=UpdateForm(request.POST)
        if form.is_valid():
            update=form.save(commit=False)
            
            update.save()
            messages.success(request, 'Record added successfully')  
            return redirect('pdf')
    context ={'form': form}
    return render(request, 'web/insert.html', context)
@login_required(login_url='/login')
@never_cache
def user_list(request):
    # =====================================================
    # USERS / MEMBERS
    # =====================================================

    users = User.objects.all()
    User_count = users.count()
    lists = room.objects.all()

    # =====================================================
    # MESSAGES / PAYMENTS
    # =====================================================

    R_messages = message.objects.all()
    payment = update.objects.all()

    # =====================================================
    # SUCCESSFUL MPESA CONTRIBUTIONS
    # =====================================================

    payments = MpesaTransaction.objects.filter(
        status="Success"
    )

    mpesa_total = (
        payments.aggregate(
            total=Sum('amount')
        )['total']
        or Decimal('0.00')
    )


    # =====================================================
    # OPENING BALANCES
    #
    # OpeningBalance.user points to the member's
    # Django User.
    # =====================================================

    opening_balance_total = (
        OpeningBalance.objects.aggregate(
            total=Sum('amount')
        )['total']
        or Decimal('0.00')
    )


    # =====================================================
    # TOTAL AMOUNT
    #
    # Opening Balance + M-Pesa Contributions
    # =====================================================

    total_amount = (
        Decimal(str(opening_balance_total))
        +
        Decimal(str(mpesa_total))
    )


    # =====================================================
    # CASH EXPENDITURE
    # =====================================================

    cash = cash_expenditure.objects.aggregate(
        total=Sum('amount')
    )

    cash_total = (
        cash['total']
        or Decimal('0.00')
    )

    cash_total = Decimal(str(cash_total))


    # =====================================================
    # CURRENT / NET BALANCE
    # =====================================================

    difference = total_amount - cash_total


    # =====================================================
    # CONTEXT
    # =====================================================

    context = {
        'users': users,
        'payments': payments,
        'difference': difference,
        # This now includes:
        # Opening Balance + M-Pesa Contributions
        'total_amount': total_amount,
        # Optional: useful if you want to display
        # the two figures separately in the dashboard
        'mpesa_total': mpesa_total,
        'opening_balance_total': opening_balance_total,
        'cash': cash_total,
        'lists': lists,
        'User_count': User_count,
        'R_messages': R_messages,
        'payment': payment,
    }

    return render(
        request,
        'web/user_list.html',
        context
    )
@require_http_methods(["GET", "POST"])
def loginPage(request):
    page= 'login'
    if request.user.is_authenticated: 
        return redirect('root')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            User= user.objects.get(username=username)
        except:
            messages.error(request, '')
            User = authenticate(request, username=username, password=password)
            if User is not None:
                login(request, User)
                return redirect('root')
            else:
              messages.error(request, 'Wrong password or username')  
    context = {'page' : page}
    return render(request,'web/registration_login.html',context)
def logoutUser(request):
    logout(request)
    return redirect('login')
@require_http_methods(["GET", "POST"])
def registerPage(request):
    form = CustomUserCreationForm()
    if request.method == 'POST':
     form = CustomUserCreationForm(request.POST)
     if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            user.save()
            login(request, user, backend='allauth.account.auth_backends.AuthenticationBackend')
            messages.success(request, 'success')  
            return redirect('createRoom')
     
     
      
            


    return render(request,'web/registration_login.html', { 'form' : form})
@login_required(login_url='/login')
@never_cache
def home(request): 
    q = request.GET.get('q') if request.GET.get('q') != None  else ''
   # tips = room.objects.filter(host=request.user) 
    tips = room.objects.filter(
        Q(group__name__icontains=q) |
        Q(FullName__icontains=q) |
        Q(HBC__icontains=q) 
        
        
                                     
                                )
    if request.user.is_authenticated:

        room_exists = room.objects.filter(host=request.user).exists()

        if not room_exists:
            return redirect("createRoom")   # url name of room form  
    
    
    topics=group.objects.all()
    contributions= update.objects.filter(user_name=request.user)
    adm = Msg.objects.all()
    users = User.objects.all()
    room_count= tips.count()
    R_messages=message.objects.all()
    context ={'tips': tips, 'contributions': contributions, 'topics': topics, 'room_count' : room_count, 'adm': adm,'users' : users, 'R_messages':R_messages}
    return render(request, 'web/home.html', context)
@require_http_methods(["GET", "POST"])
def index(request, pk):
    room_obj= room.objects.get(id=pk)
    payment= update.objects.filter(transaction_id=pk,user_name=request.user)
    #total_amount = payment.aggregate(Sum('amount'))['amount__sum'] or 0
    adm = Msg.objects.all()
    R_messages=room_obj.message_set.all().order_by('-created')

    if request.method =='POST':
        messages= message.objects.create(
            user=request.user,
            room = room_obj,
            body=request.POST.get('body')

        )
        return redirect('index', pk=index.id)
    
  
    context ={'room': room_obj, 'payment':payment,'R_messages': R_messages, 'adm':adm,}
    return render(request, 'web/index.html', context)
@require_http_methods(["GET", "POST"])
@login_required(login_url='/login')
def createRoom(request): 
    form= RoomForm()
    if request.method == 'POST':
        form=RoomForm(request.POST)
        
        if form.is_valid():
            room=form.save(commit=False)
            room.host=request.user
            room.save()
            return redirect('home')
     
         
        
        else:
              messages.error(request, 'You have an existing profile')
    context = {'form' : form}
    return render(request, 'web/room_form.html', context)
@require_http_methods(["GET", "POST"])
@login_required(login_url='/login')
def updateRoom(request, pk):
    update = room.objects.get(id=pk)
    form=RoomForm(instance=update)
    #if request.user != update.host:
        #return HttpResponse('You are not allowed here')
    if request.method == 'POST':
        form=RoomForm(request.POST, instance=update)
        if form.is_valid():
            form.save()
            return redirect('home')
    context= {'update' : update, 'form' : form}
    return render(request, 'web/room_form.html', context)
@require_http_methods(["GET", "POST"])
@login_required(login_url='/login')
def deleteRoom(request, pk):
    remove=room.objects.get(id=pk)
    if request.method == 'POST':
        remove.delete()
        return redirect('members')
    return render(request, 'web/delete.html', {'obj' :remove} )
@login_required(login_url='/login')
@user_passes_test(lambda u: u.is_superuser)
def pdf(request): 
    detail=room.objects.all()
    payment=update.objects.all() 
    Totals = update.objects.all().aggregate(total=Sum('transaction'))
    cash= cash_expenditure.objects.all().aggregate(total=Sum('Amount'))
    cash_total = cash['total'] or 0
    update_total = Totals['total'] or 0
    difference = update_total - cash_total

    return render(request, 'web/pdf.html',  {'payment': payment, 'detail': detail, 'difference':difference, 'cash': cash['total'] , 'Totals' : Totals['total']})
@login_required(login_url='/login')
def people(request, pk):
    detail=room.objects.get(id=pk)
    pays=update.objects.all() 
    context ={'detail':detail, 'pays': pays}
    return render(request, 'web/people.html', context)
@login_required(login_url='/login')
def Panel(request):
    payment=update.objects.all() 
    detail=room.objects.all()
    
    contributions = MpesaTransaction.objects.filter(status="Success")

    group = request.GET.get('group')

    selected_group = None

    if group:
        contributions = contributions.filter(profile__group__name__iexact=group)
        selected_group = group
    
    total_count = contributions.count()

    total_amount = sum(t.amount for t in contributions)

    return render(request, 'web/Panel.html', {
        'contributions': contributions,
        'total_amount': total_amount,
        'total_count': total_count,
        'selected_group': selected_group
    })
   
@login_required(login_url='/login')
def single(request):
    q = request.GET.get('q') if request.GET.get('q') != None  else ''
    payment=update.objects.filter(
        Q(choose__icontains=q) |
        Q(amount__icontains=q) |
         Q(choice__icontains=q) 
    )
    #single=update.objects.get  (id=pk)
    r_count=payment.count()
    context = {'payment': payment,'r_count':r_count }
    return render(request, 'web/import.html', context)
@login_required(login_url='/login')
@user_passes_test(lambda u: u.is_superuser)
def members(request):
    query = request.GET.get('q')
    gci =update.objects.all()
    member=room.objects.all().order_by('-created')
    users = User.objects.all().order_by('-date_joined')
   
    if query:
        member = member.filter(
            FullName__icontains=query
        ) | member.filter(
            Email__icontains=query
        ) | member.filter(
            HBC__icontains=query
        )
    context={'gci':gci, 'member':member, 'users':users, 'query':query}
    return render(request, 'web/members.html',  context)
@login_required(login_url='/login')
@user_passes_test(lambda u: u.is_superuser)
def admi(request):
    query = request.GET.get('q')
    users = User.objects.all().order_by('-date_joined')
    if query:
        users = users.filter(
            username__icontains=query
        ) | users.filter(
            email__icontains=query
        ) 
    return render(request, 'web/admin.html', {'users': users, 'query': query})
@login_required(login_url='/login')
def Message(request):
     R_messages=message.objects.all()
     adm =Msg.objects.all()
    
     
     context={ 'R_messages': R_messages, 'adm':adm}
     return render(request, 'web/Messages.html', context)
@login_required(login_url='/login')
@require_http_methods(["GET", "POST"])
def Msge(request):
    jay = Msg.objects.all()
    form= MsgForm()
    if request.method =='POST':
        form=MsgForm(request.POST)
        if form.is_valid():
            Meso=form.save(commit=False)
            
            Meso.save()
            return redirect('Message')
    context= {'form':form, 'jay': jay}
    return render(request, 'web/update_message.html', context)
@login_required(login_url='/login')
def gci_groups(request):
     q = request.GET.get('q') if request.GET.get('q') != None  else ''
     tips = room.objects.filter(
        Q(group__name__icontains=q) |
        Q(FullName__icontains=q) 
         
        
        
                                     
                                )
     topics=group.objects.all()
     
     context= {'tips':tips, 'topics':topics}

     return render(request, 'web/gci_groups.html', context)
@login_required(login_url='/login')
@require_http_methods(["GET", "POST"])
def meso(request):
    jay = message.objects.all()
    form =MessageForm()
    if request.method == 'POST':
        form=MessageForm(request.POST)
        if form.is_valid():
           meso=form.save(commit=False)   
           meso.save()
           return redirect('meso')
    context={'form': form , 'jay':jay}
    return render(request, 'web/update_message.html', context)
@login_required(login_url='/login')
@require_http_methods(["GET", "POST"])
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            # Update session hash to keep the user logged in after password change
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('password_change_done')  # Redirect to a success page
        else:
            
            messages.error(request, 'Please Check passwords.')
    else:
        form = PasswordChangeForm(user=request.user)

    return render(request, 'web/reset_password.html', {'form': form})

def root(request):
    return render(request, 'web/root.html')

def about(request):
    return render(request, 'web/about.html')

def activities(request):
    return render(request, 'web/activities.html')
@login_required(login_url='/login')
def deleteRecord(request, pk):
    cancel=update.objects.get(id=pk)
    if request.method == 'POST':
        cancel.delete()
        messages.success(request, 'Record Deleted') 
        return redirect('user_list')
    
    return render(request, 'web/delete.html', {'obj' :cancel} )

def updateRecord(request, pk):
    change = update.objects.get(id=pk)
    form=UpdateForm(instance=change)
    #if request.user != change.host:
        #return HttpResponse('You are not allowed here')
    if request.method == 'POST':
        form=UpdateForm(request.POST, instance=change)
        if form.is_valid():
            form.save()
            return redirect('pdf')
    context= {'change' : change, 'form' : form}
    return render(request, 'web/insert.html', context)



@login_required(login_url="/login")
@user_passes_test(lambda u: u.is_superuser)
def expenses(request):

    query = request.GET.get("q", "").strip()

    # =====================================================
    # GET EXPENSES
    # =====================================================

    expenses = (
        cash_expenditure.objects
        .select_related("member")
        .order_by("-date", "-created")
    )


    # =====================================================
    # SEARCH
    # =====================================================

    if query:

        expenses = expenses.filter(
            Q(member__FullName__icontains=query)
            | Q(Expense__icontains=query)
            | Q(category__icontains=query)
            | Q(Date__icontains=query)
            | Q(Amount__icontains=query)
        )


    # =====================================================
    # TOTAL EXPENSES
    # =====================================================

    total_expenses = (
        expenses.aggregate(
            total=Sum("amount")
        )["total"] or 0
    )


    # =====================================================
    # CATEGORY TOTALS
    # =====================================================

    wedding_total = (
        expenses.filter(
            category__iexact="Wedding"
        ).aggregate(
            total=Sum("amount")
        )["total"] or 0
    )


    baby_shower_total = (
        expenses.filter(
            category__iexact="Baby_Shower"
        ).aggregate(
            total=Sum("amount")
        )["total"] or 0
    )


    bereavement_total = (
        expenses.filter(
            category__iexact="Bereavement"
        ).aggregate(
            total=Sum("amount")
        )["total"] or 0
    )


    # =====================================================
    # CONTEXT
    # =====================================================

    context = {

        "expenses": expenses,

        "query": query,

        "total_expenses": total_expenses,

        "wedding_total": wedding_total,

        "baby_shower_total": baby_shower_total,

        "bereavement_total": bereavement_total,

    }


    return render(
        request,
        "web/expenses.html",
        context
    )

def more(request, pk):
      payment= update.objects.get(id=pk)
     

      context ={'payment' : payment }

      return render(request, 'web/more.html', context)

@login_required(login_url='/login')
@csrf_exempt
def mpesa(request):
    profile = room.objects.get(host=request.user)
    timestamp = timezone.localtime().strftime('%Y%m%d%H%M%S')
    form = MpesaForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        phone_number = form.cleaned_data['phone_number']
        amount = form.cleaned_data['amount']

        # Use logged-in user name (cleaner)
        full_name = request.user.get_full_name() or request.user.username

        # Convert 07XXXXXXXX → 2547XXXXXXXX
        if phone_number.startswith("0"):
            phone_number = "254" + phone_number[1:]

        cl = MpesaClient()

        try:
            response = cl.stk_push(
                phone_number=phone_number,
                amount=amount,
                account_reference='GCI WELFARE',
                transaction_desc='Welfare Contribution',
                callback_url='https://biogenetic-supergenerous-rosann.ngrok-free.dev/mpesa/callback/'
            )

            # ✅ Save transaction linked to user
            MpesaTransaction.objects.create(
                user=request.user,  # 🔑 IMPORTANT
                profile=profile,
                full_name=full_name,
                phone_number=phone_number,
                amount=amount,
                checkout_request_id=response.checkout_request_id,
                merchant_request_id=response.merchant_request_id,
                status="Pending"  # must match model choices
            )

            messages.success(request, "STK push sent. Complete payment on your phone.")
            return redirect("home")

        except Exception as e:
            print("MPESA ERROR:", e)
            messages.error(request, "Payment request failed. Try again.")
            return redirect("mpesa")

    return render(request, "web/pay.html", {
        "form": form,
        "timestamp": timestamp
    })
@csrf_exempt
def mpesa_callback(request):
    data = json.loads(request.body)

    try:
        stk_callback = data['Body']['stkCallback']

        checkout_request_id = stk_callback['CheckoutRequestID']
        result_code = stk_callback['ResultCode']

        transaction = MpesaTransaction.objects.get(
            checkout_request_id=checkout_request_id
        )

        if result_code == 0:
            transaction.status = "success"
        else:
            transaction.status = "failed"

        transaction.save()

    except Exception as e:
        print("Callback Error:", e)

    return JsonResponse({"ResultCode": 0, "ResultDesc": "Accepted"})
@login_required(login_url='/login')
@user_passes_test(lambda u: u.is_superuser)
def payment_list(request):

    # -------------------------------------------------
    # GET FILTERS
    # -------------------------------------------------

    query = request.GET.get('q', '').strip()
    transaction_date = request.GET.get('date', '').strip()

    # -------------------------------------------------
    # BASE QUERYSET
    # -------------------------------------------------

    payments = MpesaTransaction.objects.filter(
        status="Success"
    )

    # -------------------------------------------------
    # GENERAL SEARCH
    # -------------------------------------------------

    if query:

        payments = payments.filter(
            Q(full_name__icontains=query) |
            Q(phone_number__icontains=query) |
            Q(checkout_request_id__icontains=query) |
            Q(mpesa_receipt_number__icontains=query)
        )

    # -------------------------------------------------
    # MONTH FILTER
    # -------------------------------------------------

    if transaction_date:

        try:

            selected_date = datetime.strptime(
                transaction_date,
                '%Y-%m-%d'
            )

            # First day of selected month
            month_start = selected_date.replace(
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0
            )

            # First day of next month
            month_end = month_start + relativedelta(
                months=1
            )

            payments = payments.filter(
                created_at__gte=month_start,
                created_at__lt=month_end
            )

        except ValueError:
            pass

    # -------------------------------------------------
    # ORDER
    # -------------------------------------------------

    payments = payments.order_by('-created_at')

    # -------------------------------------------------
    # TOTAL
    # -------------------------------------------------

    total_amount = payments.aggregate(
        total=Sum('amount')
    )['total'] or 0

    # -------------------------------------------------
    # OTHER DATA
    # -------------------------------------------------

    check = MpesaTransaction.objects.all().order_by(
        'full_name'
    )

    # -------------------------------------------------
    # CONTEXT
    # -------------------------------------------------

    context = {
        'payments': payments,
        'query': query,
        'transaction_date': transaction_date,
        'check': check,
        'total_amount': total_amount,
    }

    return render(
        request,
        'web/payment.html',
        context
    )
    #upload excel


def download_members_excel(request):

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    # =====================================================
    # HEADERS
    # =====================================================

    ws.append([
        'ID',
        'Username',
        'Full Name',
        'Email',
        'Phone',
        'ID Number',
        'Group',
        'HBC',
        'Parent',
        'Child',
        'Spouse',
        'Updated',
        'Created',
    ])

    # =====================================================
    # FETCH MEMBERS
    # =====================================================

    members = room.objects.select_related(
        'host',
        'group'
    ).all()

    # =====================================================
    # ADD MEMBERS
    # =====================================================

    for member in members:

        # User
        username = (
            member.host.username
            if member.host
            else ""
        )

        # Group
        group_name = (
            member.group.name
            if member.group
            else ""
        )

        # HBC
        hbc = member.get_HBC_display()

        ws.append([
            member.id,
            username,
            member.FullName or "",
            member.Email or "",
            member.phone or "",
            member.id_number or "",
            group_name,
            hbc,
            member.Parent or "",
            member.Child or "",
            member.Spouse or "",
            member.updated.strftime("%Y-%m-%d %H:%M:%S")
            if member.updated
            else "",
            member.created.strftime("%Y-%m-%d %H:%M:%S")
            if member.created
            else "",
        ])

    # =====================================================
    # FORMAT HEADER
    # =====================================================

    for cell in ws[1]:

        cell.font = openpyxl.styles.Font(
            bold=True
        )

        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center",
            vertical="center"
        )

    # =====================================================
    # AUTO COLUMN WIDTH
    # =====================================================

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                cell_length = len(
                    str(cell.value)
                )

                if cell_length > max_length:
                    max_length = cell_length

            except:

                pass

        ws.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            40
        )

    # =====================================================
    # FREEZE HEADER
    # =====================================================

    ws.freeze_panes = "A2"

    # =====================================================
    # FILTER
    # =====================================================

    ws.auto_filter.ref = ws.dimensions

    # =====================================================
    # RESPONSE
    # =====================================================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="members.xlsx"'
    )

    wb.save(response)

    return response




def download_contributions_excel(request):
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pdf"

    # Headers
    ws.append(['Username', 'Full Name', 'Year', 'Month', 'Amount', 'Phone Number', 'Transaction ID'])

    # Fetch all rooms
    pdf = update.objects.all()

    for m in pdf:
        username = m.user_name.username if m.user_name else ""  # prevent None error
        ws.append([username, m.transaction.full_name, m.choice, m.choose, m.transaction.amount, m.transaction.phone_number, m.transaction.checkout_request_id])

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    wb.save(response)
    return response
def upload_cash_expenditure(request):

    if request.method == "POST":
        excel_file = request.FILES.get('file')

        if not excel_file:
            return JsonResponse({"error": "No file uploaded"})

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        headers = [str(cell.value).strip() for cell in sheet[1]]
        required = ["Expense", "Date", "Amount"]

        if headers[:3] != required:
            return JsonResponse({
        "error": f"Invalid Excel format. Found: {headers}"
    })

        count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            expense, date, amount = row

            if not expense or not amount:
                continue  # skip empty rows

            cash_expenditure.objects.create(
                Expense=expense,
                Date=str(date),
                Amount=amount
            )

            count += 1

        messages.success(request, f"{count} records uploaded successfully")
        return redirect("members")  # change to your page

    return redirect("members")
@login_required(login_url="/login")
def contributions(request, pk=None):

    # =====================================================
    # MEMBER
    # =====================================================

    member = request.user


    # =====================================================
    # SUCCESSFUL M-PESA CONTRIBUTIONS
    # =====================================================

    contributions = (
        MpesaTransaction.objects
        .filter(
            user=request.user,
            status="Success"
        )
        .order_by("-created_at")
    )


    # =====================================================
    # TOTAL M-PESA PAYMENTS
    # =====================================================

    total_paid = (
        contributions.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )

    total_paid = Decimal(
        str(total_paid)
    )


    # =====================================================
    # BALANCE BROUGHT FORWARD
    # =====================================================

    opening_balance = (
        OpeningBalance.objects
        .filter(user=request.user)
        .aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )

    opening_balance = Decimal(
        str(opening_balance)
    )


    # =====================================================
    # CURRENT CONTRIBUTION BALANCE
    # =====================================================

    current_balance = (
        opening_balance + total_paid
    )


    # =====================================================
    # WELFARE MONTHLY STATUS
    # =====================================================

    arrears = calculate_welfare_arrears(
        request.user
    )


    monthly_status = arrears.get(
        "monthly_status",
        []
    )


    paid_months_list = arrears.get(
        "paid_months_list",
        []
    )


    missing_months_list = arrears.get(
        "missing_months_list",
        []
    )


    paid_months = arrears.get(
        "paid_months",
        0
    )


    months_missing = arrears.get(
        "months_missing",
        0
    )


    amount_due = arrears.get(
        "amount_due",
        Decimal("0.00")
    )


    monthly_fee = arrears.get(
        "monthly_fee",
        Decimal("0.00")
    )


    total_required = arrears.get(
        "total_required",
        Decimal("0.00")
    )


    # =====================================================
    # WELFARE EXPENSES
    # =====================================================

    expenses = (
        cash_expenditure.objects
        .filter(
            member=request.user
        )
        .order_by(
            "-date",
            "-created"
        )
    )


    # =====================================================
    # TOTAL EXPENSES
    # =====================================================

    total_expenses = (
        expenses.aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )

    total_expenses = Decimal(
        str(total_expenses)
    )


    # =====================================================
    # WEDDING EXPENSES
    # =====================================================

    wedding_expenses = (
        expenses
        .filter(category="wedding")
        .aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )


    # =====================================================
    # BABY SHOWER EXPENSES
    # =====================================================

    baby_shower_expenses = (
        expenses
        .filter(category="baby_shower")
        .aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )


    # =====================================================
    # BEREAVEMENT EXPENSES
    # =====================================================

    bereavement_expenses = (
        expenses
        .filter(category="bereavement")
        .aggregate(
            total=Sum("amount")
        )["total"]
        or Decimal("0.00")
    )


    # =====================================================
    # NET BALANCE
    # =====================================================

    net_balance = (
        current_balance - total_expenses
    )


    # =====================================================
    # PAGE CONTEXT
    # =====================================================

    context = {

        # -------------------------------------------------
        # MEMBER
        # -------------------------------------------------

        "member": member,


        # -------------------------------------------------
        # CONTRIBUTIONS
        # -------------------------------------------------

        "contributions": contributions,

        "total_amount": total_paid,

        "mpesa_total": total_paid,


        # -------------------------------------------------
        # MONTHLY WELFARE STATUS
        # -------------------------------------------------

        "monthly_fee": monthly_fee,

        "total_required": total_required,

        "monthly_status": monthly_status,

        "paid_months_list": paid_months_list,

        "missing_months_list": missing_months_list,

        "paid_months": paid_months,

        "months_missing": months_missing,

        "amount_due": amount_due,


        # -------------------------------------------------
        # BALANCE
        # -------------------------------------------------

        "opening_balance": opening_balance,

        "current_balance": current_balance,

        "net_balance": net_balance,


        # -------------------------------------------------
        # EXPENSES
        # -------------------------------------------------

        "expenses": expenses,

        "total_expenses": total_expenses,

        "wedding_expenses": wedding_expenses,

        "baby_shower_expenses": baby_shower_expenses,

        "bereavement_expenses": bereavement_expenses,

    }


    # =====================================================
    # RENDER
    # =====================================================

    return render(
        request,
        "web/contributions.html",
        context
    )
@login_required
@user_passes_test(lambda u: u.is_superuser)
def bulk_sms_view(request):
    form = BulkMessageForm()

    if request.method == "POST":
        form = BulkMessageForm(request.POST)

        if form.is_valid():
            message = form.cleaned_data['message']

            users = MpesaTransaction.objects.filter(status="Success")

            phone_numbers = [
                user.phone_number for user in users if user.phone_number
            ]

            if not phone_numbers:
                messages.error(request, "No valid phone numbers found.")
                return redirect('bulk_sms')

            # ✅ SEND SMS AND CAPTURE RESPONSE
            response = send_bulk_sms(phone_numbers, message)

            success_count = 0
            failed_count = 0

            # ✅ CHECK RESPONSE
            if response:
                recipients = response.get('SMSMessageData', {}).get('Recipients', [])

                for r in recipients:
                    if r.get('status') == "Success":
                        success_count += 1
                    else:
                        failed_count += 1

                messages.success(
                    request,
                    f"SMS Sent: {success_count} success, {failed_count} failed"
                )
            else:
                messages.error(request, "Failed to send SMS. Check API.")

            # ✅ SAVE LOG
            BulkMessage.objects.create(
                sender=request.user,
                message=message,
                recipients_count=len(phone_numbers)
            )

            return redirect('bulk_sms')

    context = {
        "form": form,
        "count": MpesaTransaction.objects.filter(status="Success").count()
    }


    return render(request, "web/bulk_sms.html", context)

def only_superuser(user):
    return user.is_superuser


@user_passes_test(only_superuser)
def make_admin(request, user_id):
    user = get_object_or_404(User, id=user_id)

    user.is_staff = True
    user.is_superuser = True
    user.save()

    messages.success(request, f"{user.username} is now an admin.")
    return redirect('user_list')
def remove_admin(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # prevent removing yourself accidentally
    if request.user.id == user.id:
        messages.error(request, "You cannot remove your own admin rights.")
        return redirect('admin')

    user.is_superuser = False
    user.is_staff = False
    user.save()

    messages.success(request, f"{user.username} is no longer an admin.")
    return redirect('admi')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def manual_transaction(request):

    form = MpesaTransactionForm()

    if request.method == 'POST':

        form = MpesaTransactionForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect('home')

    context = {
        'form': form
    }

    return render(request, 'web/manual_transaction.html', context)
    








    

    


 
    



    






